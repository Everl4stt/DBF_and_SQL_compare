import pandas as pd
import logging
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from typing import List, Dict
from config import exclude_cols


class ResultComparator:
    def __init__(self):
        # Стили для оформления
        self.diff_fill = PatternFill(
            start_color='CD5C5C',  # Красный для различий
            end_color='CD5C5C',
            fill_type='solid'
        )
        self.dbf_fill = PatternFill(
            start_color='FFD9EAD3',  # Светло-зеленый для DBF
            end_color='FFD9EAD3',
            fill_type='solid'
        )
        self.db_fill = PatternFill(
            start_color='FFD4E6F1',  # Светло-голубой для DB
            end_color='FFD4E6F1',
            fill_type='solid'
        )
        self.header_fill = PatternFill(
            start_color='FFD3D3D3',  # Серый для заголовков
            end_color='FFD3D3D3',
            fill_type='solid'
        )
        self.header_font = Font(bold=True)
        self.logger = logging.getLogger('ResultComparator')
        self.count_compare = 0

    def compare_results(self, dbf_df: pd.DataFrame, db_df: pd.DataFrame) -> List[Dict]:
        """Сравнивает два DataFrame по строкам с одинаковыми SPN и DATO"""
        # Находим общие колонки (без учета префиксов)
        idx = 0

        common_columns = self._find_common_columns(dbf_df.columns, db_df.columns)

        if not common_columns:
            raise ValueError("Нет общих колонок для сравнения")

        comparison_results = []

        # Объединяем SPN и DATO для поиска пар
        dbf_df['COMP_KEY'] = dbf_df['SPN'].astype(str) + '_' + dbf_df['eu.DATO'].astype(str)
        db_df['COMP_KEY'] = db_df['SPN'].astype(str) + '_' + db_df['eu.DATO'].astype(str)

        # Находим общие ключи
        common_keys = set(dbf_df['COMP_KEY']).intersection(set(db_df['COMP_KEY']))

        if not common_keys:
            raise ValueError("Нет строк с одинаковыми SPN и DATO")

        for key in common_keys:
            dbf_rows = dbf_df[dbf_df['COMP_KEY'] == key]
            db_rows = db_df[db_df['COMP_KEY'] == key]
            # Для каждой пары строк с одинаковыми SPN и DATO
            for _, dbf_row in dbf_rows.iterrows():
                idx +=1
                for _, db_row in db_rows.iterrows():
                    # Добавляем заголовок с SPN и DATO
                    comparison_results.append({
                        'Type': 'Header',
                        'Field': f"{idx}. SPN: {dbf_row['SPN']}, DATO: {dbf_row['eu.DATO']}",
                        'DBF_Value': '',
                        'DB_Value': '',
                        'Status': 'MATCH'
                    })

                    # Сравниваем общие колонки
                    for col in common_columns:
                        dbf_val = dbf_row[col] if col in dbf_row else None
                        db_val = db_row[col] if col in db_row else None

                        if isinstance(dbf_val, str) and isinstance(db_val, str):
                            dbf_val = dbf_val.capitalize()
                            db_val = db_val.capitalize()

                        if dbf_val == db_val or (not dbf_val and not dbf_val) or (col in exclude_cols):
                            status = 'MATCH'
                        else:
                            status = 'DIFF'
                            self.count_compare += 1

                        comparison_results.append({
                            'Type': 'Data',
                            'Field': col,
                            'DBF_Value': dbf_val,
                            'DB_Value': db_val,
                            'Status': status
                        })

                    # Добавляем разделитель
                    comparison_results.append({
                        'Type': 'Separator',
                        'Field': '---',
                        'DBF_Value': '---',
                        'DB_Value': '---',
                        'Status': ''
                    })

        return comparison_results

    def get_count_compare(self):
        tmp = self.count_compare
        self.count_compare = 0
        return tmp

    def _find_common_columns(self, cols1: List[str], cols2: List[str]) -> List[str]:
        """Находит общие колонки"""
        # Удаляем префиксы (DBF_, DB_ и т.д.)
        base_cols1 = {col for col in cols1}
        base_cols2 = {col for col in cols2}

        return list(base_cols1 & base_cols2)

    def save_comparison(self, comparison_data: List[Dict], output_path: str):
        """Сохраняет сравнение с подсветкой различий"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison Results"

            # Заголовки
            headers = ['Field', 'DBF_Value', 'DB_Value', 'Status']
            ws.append(headers)

            # Форматируем заголовки
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font

            # Записываем данные
            for row_data in comparison_data:
                if row_data['Type'] == 'Separator':
                    ws.append(['---'] * 4)
                    continue

                row = [
                    row_data['Field'],
                    row_data['DBF_Value'],
                    row_data['DB_Value'],
                    row_data['Status']
                ]
                ws.append(row)

                # Форматируем строки
                current_row = ws.max_row

                if row_data['Type'] == 'Header':
                    for cell in ws[current_row]:
                        cell.font = self.header_font
                elif row_data['Status'] == 'DIFF':
                    for cell in ws[current_row]:
                        cell.fill = self.diff_fill
                else:
                    ws.cell(row=current_row, column=2).fill = self.dbf_fill
                    ws.cell(row=current_row, column=3).fill = self.db_fill

            for column in ws.columns:
                max_length = 0
                column_cells = [cell for cell in column]
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            wb.save(output_path)
            self.logger.info(f"Файл сравнения сохранен: {output_path}")

        except Exception as e:
            self.logger.error(f"Ошибка сохранения сравнения: {str(e)}")
            raise
